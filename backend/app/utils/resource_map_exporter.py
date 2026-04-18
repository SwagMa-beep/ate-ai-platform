"""
资源映射表Excel导出器 - 模块二
生成4个Sheet的资源映射Excel
"""
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List
from app.models.resource_map import (
    ResourceMapping, PGSConfig, PGSDetailCondition,
    PinGroupConfig, AdapterInfo
)
from app.utils.logger import setup_logger

logger = setup_logger()


def export_resource_map_excel(
    chip_name: str,
    chip_type: str,
    adapter_info: AdapterInfo,
    resource_mappings: List[ResourceMapping],
    pgs_configs: List[PGSConfig],
    pgs_details: List[PGSDetailCondition],
    pin_groups: PinGroupConfig,
    output_path: str
) -> str:
    """
    导出资源映射表Excel（4个Sheet）

    Returns:
        输出文件绝对路径
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    logger.info(f" 导出资源映射表: {chip_name}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet1: 资源映射表 ────────────────────────────────
        _write_resource_mapping_sheet(writer, resource_mappings)

        # ── Sheet2: PGS配置 ───────────────────────────────────
        _write_pgs_config_sheet(writer, pgs_configs, chip_type)

        # ── Sheet3: PGS详细条件 ───────────────────────────────
        _write_pgs_detail_sheet(writer, pgs_details)

        # ── Sheet4: 引脚分组 ──────────────────────────────────
        _write_pin_group_sheet(writer, pin_groups)

        # ── Sheet5: 适配器信息 ────────────────────────────────
        _write_adapter_info_sheet(writer, adapter_info, chip_name)

        # ── 美化 ──────────────────────────────────────────────
        _apply_styling(writer)

    logger.success(f"✅ 资源映射表已生成: {output_path}")
    return str(Path(output_path).absolute())


def _write_resource_mapping_sheet(
    writer: pd.ExcelWriter,
    mappings: List[ResourceMapping]
) -> None:
    """Sheet1: 资源映射表"""
    data = []
    for m in mappings:
        data.append({
            "引脚编号":    m.pin_no,
            "引脚名称":    m.pin_name,
            "引脚功能":    m.function,
            "方向":        m.direction,
            "STS8200S资源": m.sts_resource,
            "资源类型":    m.resource_type,
            "通道编号":    m.channel_no if m.channel_no >= 0 else "-",
            "力模式":      m.force_mode,
            "测量模式":    m.measure_mode,
            "电压量程":    m.voltage_range,
            "电流量程":    m.current_range,
            "备注":        m.notes,
        })
    pd.DataFrame(data).to_excel(
        writer, sheet_name="资源映射表", index=False
    )


def _write_pgs_config_sheet(
    writer: pd.ExcelWriter,
    configs: List[PGSConfig],
    chip_type: str
) -> None:
    """Sheet2: PGS填表配置"""
    data = []
    for c in configs:
        row = {
            "序号":         c.test_id,
            "测试项":       c.test_name,
            "PGS函数":      c.function_type,
        }

        if c.function_type == "GlobalVariable":
            row.update({
                "VECTOR_FILE":  c.vector_file,
                "AllGroup":     c.all_group,
                "INGroup":      c.in_group,
                "OutGroup":     c.out_group,
            })
        elif c.function_type in {"FIMV_PMU", "FVMI_PMU"}:
            row.update({
                "VCC_VALUE":    c.vcc_value,
                "VCC_VRANG":    c.vcc_vrang,
                "VCC_IRANG":    c.vcc_irang,
                "SELECT_GROUP": c.select_group,
                "TEST_PINS":    c.test_pins,
                "PMU_VALUE":    c.pmu_value,
                "PMU_VRANG":    c.pmu_vrang,
                "PMU_IRANG":    c.pmu_irang,
                "LIMIT_MIN":    c.limit_min,
                "LIMIT_MAX":    c.limit_max,
                "LIMIT_UNIT":   c.limit_unit,
            })
        elif c.function_type == "FUNCTION":
            row.update({
                "VCC_VALUE":    c.vcc_value,
                "VCC_VRANG":    c.vcc_vrang,
                "VCC_IRANG":    c.vcc_irang,
                "STARTLABLE":   c.start_label,
                "STOPLABLE":    c.stop_label,
                "VIH":          c.vih,
                "VIL":          c.vil,
                "VOH":          c.voh,
                "VOL":          c.vol,
            })
        elif c.function_type == "SUPPLY":
            row.update({
                "VCC_VALUE":       c.vcc_value,
                "VCC_VRANG":       c.vcc_vrang,
                "VCC_IRANG":       c.vcc_irang,
                "SelectVCC":       c.select_vcc,
                "OpenChannel":     c.open_channel,
                "OpenPinsGroup":   c.open_pins_group,
                "LIMIT_MIN":       c.limit_min,
                "LIMIT_MAX":       c.limit_max,
                "LIMIT_UNIT":      c.limit_unit,
            })
        elif c.function_type == "INLEVEL":
            row.update({
                "VCC_VALUE":       c.vcc_value,
                "STARTVoltage":    c.start_voltage,
                "STOPVoltage":     c.stop_voltage,
                "TestType":        c.test_type,
            })

        row["备注"] = c.notes
        data.append(row)

    pd.DataFrame(data).to_excel(
        writer, sheet_name="PGS配置", index=False
    )


def _write_pgs_detail_sheet(
    writer: pd.ExcelWriter,
    details: List[PGSDetailCondition]
) -> None:
    """Sheet3: PGS详细条件"""
    if not details:
        pd.DataFrame({"提示": ["暂无详细条件数据"]}).to_excel(
            writer, sheet_name="PGS详细条件", index=False
        )
        return

    data = [{
        "测试项":   d.test_name,
        "参数名":   d.param_name,
        "条件键":   d.condition_key,
        "条件值":   d.condition_value,
        "单位":     d.condition_unit,
        "说明":     d.notes,
    } for d in details]

    pd.DataFrame(data).to_excel(
        writer, sheet_name="PGS详细条件", index=False
    )


def _write_pin_group_sheet(
    writer: pd.ExcelWriter,
    groups: PinGroupConfig
) -> None:
    """Sheet4: 引脚分组"""
    data = [
        ["芯片型号",      groups.chip_name],
        ["引脚总数",      groups.pin_count],
        ["向量文件",      groups.vector_file],
        ["", ""],
        ["AllGroup（全部引脚）", ",".join(groups.all_group)],
        ["INGroup（输入引脚）",  ",".join(groups.in_group)],
        ["OutGroup（输出引脚）", ",".join(groups.out_group)],
        ["PWR引脚",              ",".join(groups.pwr_group)],
        ["GND引脚",              ",".join(groups.gnd_group)],
        ["", ""],
        ["使用说明", ""],
        ["① AllGroup", "填入PGS → GlobalVariable → AllGroup字段"],
        ["② INGroup",  "填入PGS → GlobalVariable → INGroup字段"],
        ["③ OutGroup", "填入PGS → GlobalVariable → OutGroup字段"],
        ["④ 向量文件", "ACVector Editor创建同名.vecdio文件"],
    ]
    pd.DataFrame(
        data, columns=["项目", "内容"]
    ).to_excel(writer, sheet_name="引脚分组", index=False)


def _write_adapter_info_sheet(
    writer: pd.ExcelWriter,
    adapter: AdapterInfo,
    chip_name: str
) -> None:
    """Sheet5: 适配器信息"""
    data = [
        ["适配器型号",  adapter.adapter_model],
        ["芯片类型",    adapter.chip_type],
        ["芯片插座",    adapter.socket_type],
        ["最大引脚数",  adapter.max_pin_count],
        ["", ""],
        ["可用资源",    ""],
        ["VI源通道",    ",".join(adapter.vi_channels)],
        ["DIO通道",     ",".join(adapter.dio_channels) or "无"],
        ["CBIT通道",    ",".join(adapter.cbit_channels)],
        ["TMU通道",     ",".join(adapter.tmu_channels)],
        ["", ""],
        ["使用说明",    adapter.notes],
    ]
    pd.DataFrame(
        data, columns=["项目", "说明"]
    ).to_excel(writer, sheet_name="适配器信息", index=False)


def _apply_styling(writer: pd.ExcelWriter) -> None:
    """统一Excel样式"""
    sheet_colors = {
        "资源映射表":   "1A5276",
        "PGS配置":      "145A32",
        "PGS详细条件":  "4A235B",
        "引脚分组":     "784212",
        "适配器信息":   "212F3D",
    }

    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name in writer.sheets:
        ws    = writer.sheets[sheet_name]
        color = sheet_colors.get(sheet_name, "1A5276")

        # 表头
        for cell in ws[1]:
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            cell.font      = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        # 数据行交替颜色
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            bg = "F2F3F4" if row_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.fill      = PatternFill(
                    start_color=bg, end_color=bg, fill_type="solid"
                )
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border    = border

        # 自动列宽
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            )
            ws.column_dimensions[
                col[0].column_letter
            ].width = min(max_len + 4, 50)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"