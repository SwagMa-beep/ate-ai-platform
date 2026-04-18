"""
Excel导出工具 - 面向STS8200S测试平台
支持三种测试场景的专属Sheet格式
"""
import os
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from typing import Dict
from app.utils.logger import setup_logger

logger = setup_logger()


def export_excel(
    df: pd.DataFrame,
    chip_name: str,
    output_path: str,
    chip_type: str = "UNKNOWN",
    sts_report: Dict = None,
    pin_definitions: list = None
) -> str:
    """
    导出Excel（根据chip_type选择Sheet格式）

    Args:
        df: 参数DataFrame
        chip_name: 芯片型号
        output_path: 输出路径
        chip_type: 芯片类型，决定Sheet格式
        sts_report: STS8200S适配性报告

    Returns:
        输出文件的绝对路径
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    logger.info(f" 导出Excel | 芯片: {chip_name} | 类型: {chip_type}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── 根据场景选择Sheet格式 ──────────────────────────
        if chip_type in {"DIGITAL_74", "DIGITAL_54", "DIGITAL_4000", "MEMORY"}:
            _write_digital_sheets(writer, df, chip_name)
        elif chip_type == "LDO":
            _write_ldo_sheets(writer, df, chip_name)
        elif chip_type == "EEPROM":
            _write_eeprom_sheets(writer, df, chip_name)
        else:
            _write_general_sheets(writer, df, chip_name)

        # ── 通用Sheet：STS8200S适配性报告 ──────────────────
        _write_sts_report_sheet(writer, chip_name, chip_type, sts_report)

        # ── 通用Sheet：PinMapping模板 ───────────────────────
        _write_pinmapping_sheet(writer, chip_type,pin_definitions)

        # ── 通用Sheet：Summary ──────────────────────────────
        _write_summary_sheet(writer, df, chip_name, chip_type)

        # ── 美化 ────────────────────────────────────────────
        _apply_excel_styling(writer, chip_type)

    logger.success(f"✅ Excel已生成: {output_path}")
    return str(Path(output_path).absolute())


# ============================================================
# 场景A：数字芯片Sheet
# ============================================================

def _write_digital_sheets(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    chip_name: str
) -> None:
    """数字芯片专用Sheet格式"""

    rename_map = {
        "param_name": "Test_Name",
        "condition": "Test_Condition",
        "min_val": "Min_Limit",
        "typ_val": "Typ_Value",
        "max_val": "Max_Limit",
        "unit": "Unit",
        "page": "Source_Page",
        "confidence": "Confidence",
        "sts_test_function": "STS_Function",
    }

    # Sheet1: DC_TestPlan（数字DC测试项）
    df_dc = df[
        (df.get("test_scenario", pd.Series(["GENERAL"] * len(df))) == "DIGITAL_DC")
        | df["param_name"].isin({
            "CONNECT", "FUN", "VIH", "VIL", "VOH", "VOL", "VIK",
            "II", "IIH", "IIL", "IOZH", "IOZL", "IOH", "IOL",
            "IOS", "ICCH", "ICCL", "Ron", "DeltaRon"
        })
    ].copy()

    if not df_dc.empty:
        df_dc = df_dc.rename(columns=rename_map)
        df_dc["Test_ID"] = range(1, len(df_dc) + 1)
        df_dc["Pin_List"] = ""
        df_dc["DIO_Channel"] = ""   # STS8200S DIO通道
        df_dc["VI_Channel"] = ""    # STS8200S VI源通道

        dc_cols = [
            "Test_ID", "Test_Name", "Pin_List",
            "DIO_Channel", "VI_Channel",
            "Test_Condition", "Min_Limit", "Typ_Value", "Max_Limit",
            "Unit", "STS_Function", "Source_Page", "Confidence",
            "Status", "Validation_Error", "STS_Warning"
        ]
        for c in dc_cols:
            if c not in df_dc.columns:
                df_dc[c] = ""
        df_dc[dc_cols].to_excel(
            writer, sheet_name="DC_TestPlan", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到DC测试参数"]}).to_excel(
            writer, sheet_name="DC_TestPlan", index=False
        )

    # Sheet2: AC_TestPlan（数字AC测试项）
    df_ac = df[
        df["param_name"].isin({"Tr", "tTLH", "Tf", "tTHL", "tPHL", "tPLH"})
    ].copy()

    if not df_ac.empty:
        df_ac = df_ac.rename(columns=rename_map)
        df_ac["Test_ID"] = range(1, len(df_ac) + 1)
        df_ac["Input_Pin"] = ""
        df_ac["Output_Pin"] = ""
        df_ac["Threshold_High"] = ""  # 上升沿阈值
        df_ac["Threshold_Low"] = ""   # 下降沿阈值
        df_ac["TimeSet"] = ""         # STS8200S TimeSet名称

        ac_cols = [
            "Test_ID", "Test_Name", "Input_Pin", "Output_Pin",
            "Threshold_High", "Threshold_Low",
            "Test_Condition", "Min_Limit", "Typ_Value", "Max_Limit",
            "Unit", "TimeSet", "STS_Function",
            "Source_Page", "Confidence", "Status", "Validation_Error"
        ]
        for c in ac_cols:
            if c not in df_ac.columns:
                df_ac[c] = ""
        df_ac[ac_cols].to_excel(
            writer, sheet_name="AC_TestPlan", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到AC测试参数"]}).to_excel(
            writer, sheet_name="AC_TestPlan", index=False
        )

    # Sheet3: AbsoluteMax
    df_b = df[
        (df["category"] == "B")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    _write_absolute_max_sheet(writer, df_b, rename_map)

    # Sheet4: OperatingConditions
    df_c = df[
        (df["category"] == "C")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    _write_operating_conditions_sheet(writer, df_c, rename_map)

    # Sheet5: VectorTemplate（向量文件模板说明）
    _write_vector_template_sheet(writer)

    # Sheet6: Blocked
    _write_blocked_sheet(writer, df, rename_map)


def _write_vector_template_sheet(writer: pd.ExcelWriter) -> None:
    """向量文件模板说明Sheet（STS8200S专用）"""
    vector_info = [
        ["STS8200S 测试向量文件(.vecdio)使用说明", ""],
        ["", ""],
        ["步骤", "说明"],
        ["1. 打开ACVector Editor", "双击桌面ACVector Editor图标"],
        ["2. 新建VEC文件", "文件→新建，保存类型选DIO，与工程在同一目录"],
        ["3. 创建管脚向导", "填写芯片管脚名称和数目"],
        ["4. 设置TimeSet", "设置测试时序(T1R/T1F/STBR/WAVE)"],
        ["5. 设置管脚功能", "In=输入管脚, Out=输出管脚"],
        ["6. 填写向量行", "1=高电平, 0=低电平, X=不关心"],
        ["7. 设置Label", "FUN测试: label在最后行; VOH测试: label在有效行最后"],
        ["", ""],
        ["管脚与DIO通道对应关系", ""],
        ["管脚号", "DIO通道"],
        *[[str(i), f"D{i-1}"] for i in range(1, 25)],
        ["", ""],
        ["注意事项", ""],
        ["① 新建向量行时多建2行，最后两行不作修改", ""],
        ["② 每组8路DIO共用GND，注意拨码开关位置", ""],
        ["③ 时序参数：T1R(上升边沿), T1F(下降边沿), STBR(采样点)", ""],
    ]
    pd.DataFrame(vector_info, columns=["项目", "说明"]).to_excel(
        writer, sheet_name="VectorTemplate", index=False
    )


# ============================================================
# 场景B1：LDO Sheet
# ============================================================

def _write_ldo_sheets(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    chip_name: str
) -> None:
    """LDO专用Sheet格式"""

    rename_map = {
        "param_name": "Test_Name",
        "condition": "Test_Condition",
        "min_val": "Min_Limit",
        "typ_val": "Typ_Value",
        "max_val": "Max_Limit",
        "unit": "Unit",
        "page": "Source_Page",
        "confidence": "Confidence",
        "sts_test_function": "STS_Function",
    }

    # Sheet1: LDO_TestPlan
    ldo_params = ["VO", "Sv", "Si", "Iq"]
    df_ldo = df[
        df["param_name"].isin(ldo_params)
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()

    if not df_ldo.empty:
        df_ldo = df_ldo.rename(columns=rename_map)
        df_ldo["Test_ID"] = range(1, len(df_ldo) + 1)
        df_ldo["VI_Channel_IN"] = "FVI0"    # 输入端
        df_ldo["VI_Channel_OUT"] = "FVI1"   # 输出端
        df_ldo["Force_Mode"] = ""           # 力电压/力电流
        df_ldo["Measure_Mode"] = ""         # 测电压/测电流
        df_ldo["Test_Description"] = df_ldo["Test_Name"].map({
            "VO": "输出电压测试 - 在标准条件下测量VOUT",
            "Sv": "线性调整率 - VIN从Min扫描到Max，测ΔVOUT",
            "Si": "负载调整率 - IOUT从Min扫描到Max，测ΔVOUT",
            "Iq": "静态电流 - 无负载时测量VCC消耗电流"
        }).fillna("")

        ldo_cols = [
            "Test_ID", "Test_Name", "Test_Description",
            "VI_Channel_IN", "VI_Channel_OUT",
            "Force_Mode", "Measure_Mode",
            "Test_Condition", "Min_Limit", "Typ_Value", "Max_Limit",
            "Unit", "STS_Function",
            "Source_Page", "Confidence", "Status",
            "Validation_Error", "STS_Warning"
        ]
        for c in ldo_cols:
            if c not in df_ldo.columns:
                df_ldo[c] = ""
        df_ldo[ldo_cols].to_excel(
            writer, sheet_name="LDO_TestPlan", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到LDO测试参数"]}).to_excel(
            writer, sheet_name="LDO_TestPlan", index=False
        )

    # Sheet2: LDO_TestMethod（测试方法说明）
    method_data = [
        ["参数", "测试方法", "STS8200S实现", "力源", "测量"],
        [
            "VO(输出电压)",
            "固定VIN，固定IOUT，测VOUT",
            "FOVI_Test: Force V, Measure I → 换算",
            "FVI0: Force VIN", "FVI1: Measure VOUT"
        ],
        [
            "Sv(电压调整率)",
            "固定IOUT，VIN从Min扫到Max，测ΔVOUT",
            "FOVI_Test: 循环改变VIN，记录VOUT",
            "FVI0: Sweep VIN", "FVI1: Measure VOUT"
        ],
        [
            "Si(负载调整率)",
            "固定VIN，IOUT从Min扫到Max，测ΔVOUT",
            "FOVI_Test: 循环改变IOUT，记录VOUT",
            "FVI0: Force VIN", "FVI1: Sweep IOUT"
        ],
        [
            "Iq(静态电流)",
            "固定VIN，IOUT=0，测VCC电流",
            "QTMU_Test: 精密电流测量",
            "FVI0: Force VIN", "QTMU: Measure ICC"
        ],
    ]
    pd.DataFrame(
        method_data[1:], columns=method_data[0]
    ).to_excel(writer, sheet_name="LDO_TestMethod", index=False)

    # Sheet3: AbsoluteMax
    df_b = df[
        (df["category"] == "B")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    _write_absolute_max_sheet(writer, df_b, rename_map)

    # Sheet4: OperatingConditions
    df_c = df[
        (df["category"] == "C")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    _write_operating_conditions_sheet(writer, df_c, rename_map)

    # Sheet5: Blocked
    _write_blocked_sheet(writer, df, rename_map)


# ============================================================
# 场景B2：EEPROM Sheet
# ============================================================

def _write_eeprom_sheets(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    chip_name: str
) -> None:
    """EEPROM专用Sheet格式"""

    rename_map = {
        "param_name": "Test_Name",
        "condition": "Test_Condition",
        "min_val": "Min_Limit",
        "typ_val": "Typ_Value",
        "max_val": "Max_Limit",
        "unit": "Unit",
        "page": "Source_Page",
        "confidence": "Confidence",
        "sts_test_function": "STS_Function",
    }

    # Sheet1: EEPROM_FuncTest（功能测试项）
    func_test_data = [
        {
            "Test_ID": 1, "Test_Name": "WRITE_READ_55",
            "Test_Description": "写入0x55, 读出验证",
            "Test_Data": "0x55 (01010101B)",
            "SCL_Channel": "D0", "SDA_Channel": "D1",
            "VCC_Level": "5.0V", "STS_Function": "DIO_Test",
            "Expected": "读出数据=0x55", "Status": "待复核"
        },
        {
            "Test_ID": 2, "Test_Name": "WRITE_READ_AA",
            "Test_Description": "写入0xAA, 读出验证",
            "Test_Data": "0xAA (10101010B)",
            "SCL_Channel": "D0", "SDA_Channel": "D1",
            "VCC_Level": "5.0V", "STS_Function": "DIO_Test",
            "Expected": "读出数据=0xAA", "Status": "待复核"
        },
        {
            "Test_ID": 3, "Test_Name": "WRITE_READ_DIFF",
            "Test_Description": "写入不同数据模式, 读出验证",
            "Test_Data": "0x00/0xFF/0xA5等不同模式",
            "SCL_Channel": "D0", "SDA_Channel": "D1",
            "VCC_Level": "5.0V", "STS_Function": "DIO_Test",
            "Expected": "读出数据与写入完全一致", "Status": "待复核"
        },
    ]
    pd.DataFrame(func_test_data).to_excel(
        writer, sheet_name="EEPROM_FuncTest", index=False
    )

    # Sheet2: EEPROM_DCParam（电气参数）
    df_dc = df[
        (df["category"] == "A")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    if not df_dc.empty:
        df_dc = df_dc.rename(columns=rename_map)
        df_dc["Test_ID"] = range(1, len(df_dc) + 1)
        dc_cols = [
            "Test_ID", "Test_Name", "Test_Condition",
            "Min_Limit", "Typ_Value", "Max_Limit", "Unit",
            "STS_Function", "Source_Page", "Confidence",
            "Status", "Validation_Error", "STS_Warning"
        ]
        for c in dc_cols:
            if c not in df_dc.columns:
                df_dc[c] = ""
        df_dc[dc_cols].to_excel(
            writer, sheet_name="EEPROM_DCParam", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到EEPROM电气参数"]}).to_excel(
            writer, sheet_name="EEPROM_DCParam", index=False
        )

    # Sheet3: I2C_Timing（I2C时序参数）
    timing_params = {
        "fSCL", "tSU", "tHD", "tLOW", "tHIGH",
        "tAA", "tBUF", "tSP", "tWR", "tRC"
    }
    df_timing = df[df["param_name"].isin(timing_params)].copy()
    if not df_timing.empty:
        df_timing = df_timing.rename(columns=rename_map)
        df_timing["TimeSet_Config"] = ""  # 对应STS8200S TimeSet配置
        df_timing.to_excel(
            writer, sheet_name="I2C_Timing", index=False
        )
    else:
        # 生成AT24C01标准时序参考表
        timing_ref = [
            ["参数", "描述", "Min", "Typ", "Max", "Unit", "STS TimeSet对应"],
            ["fSCL", "SCL时钟频率", "0", "-", "400", "kHz", "1/fSCL=T1R+T1F"],
            ["tLOW", "SCL低电平时间", "1.3", "-", "-", "μs", "T1F"],
            ["tHIGH", "SCL高电平时间", "0.6", "-", "-", "μs", "T1R"],
            ["tSU:DAT", "数据建立时间", "100", "-", "-", "ns", "STBR"],
            ["tHD:DAT", "数据保持时间", "0", "-", "900", "ns", "-"],
        ]
        pd.DataFrame(
            timing_ref[1:], columns=timing_ref[0]
        ).to_excel(writer, sheet_name="I2C_Timing", index=False)

    # Sheet4: AbsoluteMax
    df_b = df[
        (df["category"] == "B")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    _write_absolute_max_sheet(writer, df_b, rename_map)

    # Sheet5: Blocked
    _write_blocked_sheet(writer, df, rename_map)


# ============================================================
# 场景C：通用模拟芯片Sheet（保持原有逻辑）
# ============================================================

def _write_general_sheets(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    chip_name: str
) -> None:
    """通用Sheet格式（保持原有逻辑）"""

    rename_map = {
        "param_name": "Test_Name",
        "condition": "Condition",
        "min_val": "Min_Limit",
        "typ_val": "Typ_Value",
        "max_val": "Max_Limit",
        "unit": "Unit",
        "page": "Source_Page",
        "confidence": "Confidence",
    }

    df_a = df[
        (df["category"] == "A")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    df_b = df[
        (df["category"] == "B")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()
    df_c = df[
        (df["category"] == "C")
        & (~df["Status"].str.contains("已拦截", na=False))
    ].copy()

    # TestPlan Sheet
    if not df_a.empty:
        df_out = df_a.rename(columns=rename_map).copy()
        df_out["Test_ID"] = range(1, len(df_out) + 1)
        df_out["Test_Type"] = "DC"
        df_out["STS_Function"] = df_a.get(
            "sts_test_function", ""
        )
        cols = [
            "Test_ID", "Test_Name", "Test_Type",
            "Condition", "Min_Limit", "Typ_Value", "Max_Limit",
            "Unit", "STS_Function", "Source_Page", "Confidence",
            "Status", "Validation_Error", "STS_Warning"
        ]
        for c in cols:
            if c not in df_out.columns:
                df_out[c] = ""
        df_out[cols].to_excel(
            writer, sheet_name="TestPlan", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到A类参数"]}).to_excel(
            writer, sheet_name="TestPlan", index=False
        )

    _write_absolute_max_sheet(writer, df_b, rename_map)
    _write_operating_conditions_sheet(writer, df_c, rename_map)
    _write_blocked_sheet(writer, df, rename_map)


# ============================================================
# 通用Sheet辅助函数
# ============================================================

def _write_absolute_max_sheet(
    writer: pd.ExcelWriter,
    df_b: pd.DataFrame,
    rename_map: Dict
) -> None:
    """写入绝对最大值Sheet"""
    if not df_b.empty:
        df_out = df_b.rename(columns=rename_map).copy()
        df_out["ID"] = range(1, len(df_out) + 1)
        df_out["Safety_Note"] = "测试时不可超过此值"
        cols = [
            "ID", "Test_Name", "Condition",
            "Min_Limit", "Typ_Value", "Max_Limit",
            "Unit", "Safety_Note", "Source_Page",
            "Confidence", "Status", "Validation_Error"
        ]
        # 列名兼容
        if "Test_Condition" in df_out.columns and "Condition" not in df_out.columns:
            df_out["Condition"] = df_out["Test_Condition"]
        for c in cols:
            if c not in df_out.columns:
                df_out[c] = ""
        df_out[cols].to_excel(
            writer, sheet_name="AbsoluteMax", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到B类绝对最大值"]}).to_excel(
            writer, sheet_name="AbsoluteMax", index=False
        )


def _write_operating_conditions_sheet(
    writer: pd.ExcelWriter,
    df_c: pd.DataFrame,
    rename_map: Dict
) -> None:
    """写入工作条件Sheet"""
    if not df_c.empty:
        df_out = df_c.rename(columns=rename_map).copy()
        df_out["ID"] = range(1, len(df_out) + 1)
        cols = [
            "ID", "Test_Name", "Min_Limit", "Max_Limit",
            "Unit", "Source_Page", "Confidence",
            "Status", "Validation_Error"
        ]
        for c in cols:
            if c not in df_out.columns:
                df_out[c] = ""
        df_out[cols].to_excel(
            writer, sheet_name="OperatingConditions", index=False
        )
    else:
        pd.DataFrame({"提示": ["未提取到C类工作条件"]}).to_excel(
            writer, sheet_name="OperatingConditions", index=False
        )


def _write_pinmapping_sheet(
    writer: pd.ExcelWriter,
    chip_type: str,
    pin_definitions: list = None    # ← 新增参数
) -> None:
    """写入PinMapping Sheet - 支持自动填充引脚定义"""

    # ── 有引脚定义：自动填充 ──────────────────────────────
    if pin_definitions and len(pin_definitions) > 0:
        logger.info(
            f"✅ PinMapping自动填充 {len(pin_definitions)} 个引脚"
        )
        data = []
        for pin in pin_definitions:
            # 基础列（所有芯片类型都有）
            row = {
                "Pin_No":      pin.pin_no,
                "Pin_Name":    pin.pin_name,
                "Function":    pin.function,
                "Direction":   pin.direction,
                "Voltage_Max": pin.voltage_max if pin.voltage_max else "",
                "Notes":       pin.notes,
            }
            # 根据芯片类型添加资源列（模块二自动填充）
            if chip_type in {
                "DIGITAL_74", "DIGITAL_54",
                "DIGITAL_4000", "MEMORY"
            }:
                row["DIO_Channel"] = ""   # 模块二填充
                row["VI_Channel"]  = ""
                row["VIH"]         = ""
                row["VIL"]         = ""
            elif chip_type == "LDO":
                row["VI_Channel"]  = ""   # 模块二填充
                row["Force_Mode"]  = ""
                row["Measure_Mode"]= ""
            elif chip_type == "EEPROM":
                row["DIO_Channel"] = ""   # 模块二填充
                row["VI_Channel"]  = ""
                row["I2C_Role"]    = ""
            else:
                row["STS_Resource"] = ""
                row["Force_Mode"]   = ""

            data.append(row)

        pd.DataFrame(data).to_excel(
            writer, sheet_name="PinMapping", index=False
        )

    # ── 没有引脚定义：生成空模板 ─────────────────────────
    else:
        logger.warning(
            "⚠️ 未提取到引脚定义，PinMapping为空模板，请手动填写"
        )
        if chip_type in {
            "DIGITAL_74", "DIGITAL_54", "DIGITAL_4000", "MEMORY"
        }:
            cols = [
                "Pin_No", "Pin_Name", "Function", "Direction",
                "DIO_Channel", "VI_Channel", "VIH", "VIL", "Notes"
            ]
        elif chip_type == "LDO":
            cols = [
                "Pin_No", "Pin_Name", "Function", "Direction",
                "VI_Channel", "Force_Mode", "Measure_Mode", "Notes"
            ]
        elif chip_type == "EEPROM":
            cols = [
                "Pin_No", "Pin_Name", "Function", "Direction",
                "DIO_Channel", "VI_Channel", "I2C_Role", "Notes"
            ]
        else:
            cols = [
                "Pin_No", "Pin_Name", "Function", "Direction",
                "Voltage_Max", "Current_Max", "Notes"
            ]

        pd.DataFrame(columns=cols).to_excel(
            writer, sheet_name="PinMapping", index=False
        )

def _write_blocked_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    rename_map: Dict
) -> None:
    """写入被拦截参数Sheet"""
    df_blocked = df[
        df["Status"].str.contains("已拦截", na=False)
    ].copy()

    if not df_blocked.empty:
        df_blocked = df_blocked.rename(columns=rename_map).copy()
        df_blocked["ID"] = range(1, len(df_blocked) + 1)
        df_blocked.to_excel(
            writer, sheet_name="Blocked", index=False
        )
    else:
        pd.DataFrame({"提示": ["无被拦截参数"]}).to_excel(
            writer, sheet_name="Blocked", index=False
        )


def _write_sts_report_sheet(
    writer: pd.ExcelWriter,
    chip_name: str,
    chip_type: str,
    sts_report: Dict = None
) -> None:
    """写入STS8200S适配性报告Sheet"""
    report_data = [
        ["STS8200S 硬件适配性报告", ""],
        ["芯片型号", chip_name],
        ["识别芯片类型", chip_type],
        ["", ""],
        ["硬件资源限制", ""],
        ["VI源最大电压", "10.0 V"],
        ["VI源最大电流", "200 mA"],
        ["DIO通道数量", "24路 (D0~D23)"],
        ["DIO电压范围", "-0.5V ~ 5.5V"],
        ["CBIT继电器路数", "40路"],
        ["VI源分组", "左侧FVI0-3 / 右侧FVI4-7，每组共用低端"],
        ["", ""],
    ]

    if sts_report:
        report_data.append(["兼容性检查结果", ""])
        report_data.append([
            "整体兼容性",
            "✅ 兼容" if sts_report.get("is_compatible") else "❌ 存在问题"
        ])
        for issue in sts_report.get("issues", []):
            report_data.append(["⚠️ 问题", issue])
        report_data.append(["", ""])
        report_data.append(["接线建议", ""])
        for rec in sts_report.get("recommendations", []):
            report_data.append(["✓", rec])

    pd.DataFrame(
        report_data, columns=["项目", "说明"]
    ).to_excel(writer, sheet_name="STS_Report", index=False)


def _write_summary_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    chip_name: str,
    chip_type: str
) -> None:
    """写入汇总信息Sheet"""
    df_blocked = df[df["Status"].str.contains("已拦截", na=False)]
    df_valid = df[~df["Status"].str.contains("已拦截", na=False)]
    df_a = df_valid[df_valid["category"] == "A"]
    df_b = df_valid[df_valid["category"] == "B"]
    df_c = df_valid[df_valid["category"] == "C"]

    summary_data = [
        ["芯片型号", chip_name],
        ["芯片类型", chip_type],
        ["", ""],
        ["提取统计", ""],
        ["A类(电气特性/测试项)", f"{len(df_a)} 条"],
        ["B类(绝对最大值)", f"{len(df_b)} 条"],
        ["C类(工作条件)", f"{len(df_c)} 条"],
        ["已拦截(无效参数)", f"{len(df_blocked)} 条"],
        ["总计", f"{len(df)} 条"],
        ["", ""],
        ["各Sheet用途说明", ""],
    ]

    # 根据芯片类型添加对应说明
    if chip_type in {"DIGITAL_74", "DIGITAL_54", "DIGITAL_4000", "MEMORY"}:
        summary_data.extend([
            ["DC_TestPlan", "数字芯片DC测试项，直接导入STS8200S编写测试程序"],
            ["AC_TestPlan", "数字芯片AC测试项，时序参数"],
            ["VectorTemplate", "STS8200S向量文件(.vecdio)使用说明"],
            ["AbsoluteMax", "设置VI源/PMU保护限值，防止烧芯片"],
            ["OperatingConditions", "设置测试全局条件(VCC/温度)"],
            ["PinMapping", "填写引脚→DIO通道映射，供模块②使用"],
            ["STS_Report", "STS8200S硬件适配性检查报告"],
        ])
    elif chip_type == "LDO":
        summary_data.extend([
            ["LDO_TestPlan", "LDO测试项(VO/Sv/Si/Iq)，含VI源通道分配"],
            ["LDO_TestMethod", "各参数测试方法和STS8200S实现方式"],
            ["AbsoluteMax", "LDO绝对最大值，设置VI源保护限"],
            ["OperatingConditions", "LDO推荐工作条件"],
            ["PinMapping", "LDO引脚→VI源通道映射，供模块②使用"],
            ["STS_Report", "STS8200S硬件适配性检查报告"],
        ])
    elif chip_type == "EEPROM":
        summary_data.extend([
            ["EEPROM_FuncTest", "EEPROM功能测试项(读写55/AA/不同数据)"],
            ["EEPROM_DCParam", "EEPROM电气参数测试项"],
            ["I2C_Timing", "I2C时序参数，对应STS8200S TimeSet配置"],
            ["AbsoluteMax", "EEPROM绝对最大值"],
            ["PinMapping", "EEPROM引脚→DIO通道映射"],
            ["STS_Report", "STS8200S硬件适配性检查报告"],
        ])

    pd.DataFrame(
        summary_data, columns=["项目", "说明"]
    ).to_excel(writer, sheet_name="Summary", index=False)


# ============================================================
# Excel样式
# ============================================================

def _apply_excel_styling(
    writer: pd.ExcelWriter,
    chip_type: str = "UNKNOWN"
) -> None:
    """应用Excel样式（按场景设置颜色主题）"""

    # 根据芯片类型设置颜色主题
    if chip_type in {"DIGITAL_74", "DIGITAL_54", "DIGITAL_4000", "MEMORY"}:
        primary_color = "1F4E79"    # 深蓝（数字）
        secondary_color = "2E75B6"
    elif chip_type == "LDO":
        primary_color = "833C00"    # 深橙（模拟）
        secondary_color = "C55A11"
    elif chip_type == "EEPROM":
        primary_color = "375623"    # 深绿（存储）
        secondary_color = "548235"
    else:
        primary_color = "2F5496"
        secondary_color = "4472C4"

    sheet_colors = {
        # 场景A
        "DC_TestPlan": primary_color,
        "AC_TestPlan": secondary_color,
        "VectorTemplate": "7F7F7F",
        # 场景B LDO
        "LDO_TestPlan": primary_color,
        "LDO_TestMethod": secondary_color,
        # 场景B EEPROM
        "EEPROM_FuncTest": primary_color,
        "EEPROM_DCParam": secondary_color,
        "I2C_Timing": "375623",
        # 通用场景C
        "TestPlan": primary_color,
        # 通用Sheet
        "AbsoluteMax": "C00000",
        "OperatingConditions": "548235",
        "PinMapping": "7F7F7F",
        "Blocked": "808080",
        "STS_Report": "4472C4",
        "Summary": "BF8F00",
    }

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        color = sheet_colors.get(sheet_name, primary_color)

        # 表头样式
        header_fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # 自动列宽
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            )
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(
                max_len + 4, 45
            )

        # 冻结首行
        ws.freeze_panes = "A2"