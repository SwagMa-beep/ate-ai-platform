"""
STS8200S专用Excel导出器
按测试类型分类，符合实际使用习惯
"""
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List
from app.models.testplan import STSTestPlan, STSTestParam, TestType
from app.utils.logger import setup_logger

logger = setup_logger()


def export_sts_testplan(test_plan: STSTestPlan, output_path: str) -> str:
    """
    导出STS8200S测试计划Excel

    Sheet布局：
    1. Summary - 测试概览
    2. DC_Tests - DC参数测试
    3. AC_Tests - AC参数测试
    4. Functional_Tests - 功能测试
    5. Connect_Test - 连接性测试
    6. PinMapping - 引脚映射（空模板）
    """
    logger.info(f" 开始导出STS测试计划: {output_path}")

    # 按测试类型分组
    params_by_type = {
        "DC": [],
        "AC": [],
        "FUNCTION": [],
        "CONNECT": []
    }

    for param in test_plan.test_params:
        test_type = param.test_type.value
        if test_type in params_by_type:
            params_by_type[test_type].append(param)

    # 创建Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ========== Sheet 1: Summary ==========
        summary_data = [
            ["芯片型号", test_plan.chip_name],
            ["芯片类型", test_plan.chip_type.value],
            ["封装类型", test_plan.package_type],
            ["总测试项", len(test_plan.test_params)],
            ["", ""],
            ["测试类型分布", "数量"],
            ["DC参数测试", len(params_by_type["DC"])],
            ["AC参数测试", len(params_by_type["AC"])],
            ["功能测试", len(params_by_type["FUNCTION"])],
            ["连接性测试", len(params_by_type["CONNECT"])],
            ["", ""],
            ["DC参数细分", ""],
        ]

        # 添加DC参数统计
        stats = test_plan.get_stats()
        for dc_type, count in stats.get("dc_params", {}).items():
            summary_data.append([f"  {dc_type}", count])

        summary_data.append(["", ""])
        summary_data.append(["AC参数细分", ""])
        for ac_type, count in stats.get("ac_params", {}).items():
            summary_data.append([f"  {ac_type}", count])

        df_summary = pd.DataFrame(summary_data, columns=["项目", "值"])
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        # ========== Sheet 2: DC_Tests ==========
        if params_by_type["DC"]:
            dc_data = []
            for idx, param in enumerate(params_by_type["DC"], 1):
                dc_data.append({
                    "Test_ID": idx,
                    "Test_Name": param.param_name,
                    "Symbol": param.param_symbol,
                    "Sub_Type": param.dc_param_type.value if param.dc_param_type else "",
                    "Condition": param.condition,
                    "Min": param.min_val,
                    "Typ": param.typ_val,
                    "Max": param.max_val,
                    "Unit": param.unit,
                    "Page": param.page,
                    "Confidence": f"{param.confidence:.0%}",
                    "Notes": param.notes
                })

            df_dc = pd.DataFrame(dc_data)
            df_dc.to_excel(writer, sheet_name="DC_Tests", index=False)
        else:
            pd.DataFrame({"提示": ["未找到DC参数"]}).to_excel(
                writer, sheet_name="DC_Tests", index=False
            )

        # ========== Sheet 3: AC_Tests ==========
        if params_by_type["AC"]:
            ac_data = []
            for idx, param in enumerate(params_by_type["AC"], 1):
                ac_data.append({
                    "Test_ID": idx,
                    "Test_Name": param.param_name,
                    "Symbol": param.param_symbol,
                    "Sub_Type": param.ac_param_type.value if param.ac_param_type else "",
                    "Condition": param.condition,
                    "Min": param.min_val,
                    "Typ": param.typ_val,
                    "Max": param.max_val,
                    "Unit": param.unit,
                    "Page": param.page,
                    "Confidence": f"{param.confidence:.0%}",
                    "Notes": param.notes
                })

            df_ac = pd.DataFrame(ac_data)
            df_ac.to_excel(writer, sheet_name="AC_Tests", index=False)
        else:
            pd.DataFrame({"提示": ["未找到AC参数"]}).to_excel(
                writer, sheet_name="AC_Tests", index=False
            )

        # ========== Sheet 4: Functional_Tests ==========
        if params_by_type["FUNCTION"]:
            func_data = []
            for idx, param in enumerate(params_by_type["FUNCTION"], 1):
                func_data.append({
                    "Test_ID": idx,
                    "Test_Name": param.param_name,
                    "Condition": param.condition,
                    "Expected_Result": param.notes,
                    "Page": param.page
                })

            df_func = pd.DataFrame(func_data)
            df_func.to_excel(writer, sheet_name="Functional_Tests", index=False)
        else:
            pd.DataFrame({"提示": ["未找到功能测试项"]}).to_excel(
                writer, sheet_name="Functional_Tests", index=False
            )

        # ========== Sheet 5: Connect_Test ==========
        pd.DataFrame({
            "说明": [
                "连接性测试(CONNECT)用于验证DUT与测试机台的连接",
                "通常测试所有数字引脚的上拉/下拉电阻",
                "该测试项一般在Datasheet中不明确列出",
                "需要根据芯片引脚定义手动配置"
            ]
        }).to_excel(writer, sheet_name="Connect_Test", index=False)

        # ========== Sheet 6: PinMapping ==========
        pin_template = pd.DataFrame(columns=[
            "Pin_No",  # 引脚编号
            "Pin_Name",  # 引脚名称
            "Function",  # 功能（IN/OUT/PWR/GND）
            "ATE_Resource",  # ATE资源（DIO/VI）
            "ATE_Channel",  # 通道号（D0-D23 / VI0-VI7）
            "Notes"  # 备注
        ])
        pin_template.to_excel(writer, sheet_name="PinMapping", index=False)

        # ========== Sheet 7: 向量文件模板 ==========
        vector_guide = pd.DataFrame({
            "步骤": [
                "1", "2", "3", "4", "5"
            ],
            "操作": [
                "根据PinMapping填写引脚到通道的映射",
                "打开ACVector Editor软件",
                "创建.vecdio文件，设置引脚名称和时间集",
                "编写功能测试向量",
                "保存到工程目录，供测试程序调用"
            ],
            "参考": [
                "见PinMapping表",
                "软件路径: C:\\STS8200S\\ACVectorEditor.exe",
                "手册第3章",
                "FUN测试需要完整向量表",
                "文件名格式: <芯片型号>.vecdio"
            ]
        })
        vector_guide.to_excel(writer, sheet_name="向量文件指南", index=False)

        # ========== 美化样式 ==========
        _apply_sts_styling(writer)

    logger.success(f"✅ Excel导出完成: {output_path}")
    return output_path


def _apply_sts_styling(writer):
    """应用STS专用样式"""

    # 颜色方案
    colors = {
        "Summary": "FFD700",  # 金色
        "DC_Tests": "4472C4",  # 蓝色
        "AC_Tests": "70AD47",  # 绿色
        "Functional_Tests": "FFC000",  # 橙色
        "Connect_Test": "5B9BD5",  # 浅蓝
        "PinMapping": "A5A5A5",  # 灰色
        "向量文件指南": "C55A11"  # 棕色
    }

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        color = colors.get(sheet_name, "4472C4")

        # 表头样式
        header_fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
            size=11,
            name="微软雅黑"
        )

        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用到第一行
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = thin_border

        # 自动列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 数据行样式（交替颜色）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill_color = "F2F2F2" if row_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )
                cell.border = thin_border