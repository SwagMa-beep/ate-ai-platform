"""
BOM清单生成器 - 模块二
基于适配器BOM数据生成Excel清单
"""
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from typing import List, Dict
from app.utils.logger import setup_logger

logger = setup_logger()


def generate_bom_excel(
    bom_items: List[Dict],
    chip_name: str,
    adapter_model: str,
    output_path: str
) -> str:
    """
    生成BOM清单Excel

    Args:
        bom_items     : BOM数据列表
        chip_name     : 芯片型号
        adapter_model : 适配器型号
        output_path   : 输出路径

    Returns:
        输出文件绝对路径
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    logger.info(f" 生成BOM清单: {adapter_model}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet1: BOM主表 ──────────────────────────────────
        df_bom = pd.DataFrame(bom_items)
        # 添加采购说明列
        df_bom["采购注意事项"] = df_bom.apply(
            _get_procurement_note, axis=1
        )
        df_bom.to_excel(writer, sheet_name="BOM清单", index=False)

        # ── Sheet2: 统计汇总 ─────────────────────────────────
        summary = _generate_summary(bom_items, chip_name, adapter_model)
        pd.DataFrame(
            summary, columns=["项目", "说明"]
        ).to_excel(writer, sheet_name="汇总信息", index=False)

        # ── Sheet3: 关键元件说明 ─────────────────────────────
        key_parts = _generate_key_parts_sheet(adapter_model)
        pd.DataFrame(
            key_parts[1:], columns=key_parts[0]
        ).to_excel(writer, sheet_name="关键元件说明", index=False)

        # ── 样式 ─────────────────────────────────────────────
        _apply_bom_styling(writer)

    logger.success(f"✅ BOM清单已生成: {output_path}")
    return str(Path(output_path).absolute())


def _get_procurement_note(row: pd.Series) -> str:
    """根据元件类型生成采购注意事项"""
    val  = str(row.get("值",   "")).upper()
    desc = str(row.get("说明", "")).upper()

    if "AGQ200A4H" in val:
        return "松下继电器，SPST型，5V线圈，注意工作温度-40~85℃"
    elif "BUF634" in val:
        return "TI BUF634，250mA高速缓冲，注意散热，可用BUF634T(SMD)"
    elif "EL7156" in val:
        return "Microchip EL7156，高性能单管脚驱动，SOT23-5封装"
    elif "BAV99" in val:
        return "双向肖特基二极管，SOT23，注意极性方向"
    elif "ZIF14" in val or "ZIF" in val:
        return "14引脚零插入力锁紧座，确认芯片引脚间距2.54mm"
    elif "SOP8" in val:
        return "SOP8插座，确认引脚间距1.27mm，适配目标芯片封装"
    elif "0.1UF" in val:
        return "去耦电容，X5R/X7R介质，耐压≥10V"
    elif "4.7UF" in val or "10UF" in val:
        return "电解或钽电容，耐压≥16V，注意极性"
    elif "51" in val and "Ω" not in val.replace("Ω",""):
        return "精密电阻，1%精度，用于BUF634输出匹配"
    else:
        return "标准元件，按参数采购"


def _generate_summary(
    bom_items: List[Dict],
    chip_name: str,
    adapter_model: str
) -> List[List]:
    total_qty = sum(item.get("数量", 0) for item in bom_items)
    return [
        ["适配器型号",  adapter_model],
        ["适用芯片",    chip_name],
        ["元件种类数",  str(len(bom_items))],
        ["元件总数量",  str(total_qty)],
        ["", ""],
        ["采购建议",    ""],
        ["继电器",      "AGQ200A4H，建议备货×2倍数量"],
        ["IC插座",      "按芯片封装选型，建议备货×3"],
        ["去耦电容",    "0.1uF批量采购，建议备货×10"],
        ["关键IC",      "BUF634/EL7156提前确认货期"],
    ]


def _generate_key_parts_sheet(adapter_model: str) -> List:
    """关键元件详细说明"""
    header = ["元件", "型号", "功能", "关键参数", "替代型号"]
    rows   = []

    if "8281-13" in adapter_model:  # 数字适配器
        rows = [
            ["继电器K1/K2", "AGQ200A4H", "VDD/VSS供电切换",
             "SPST, 5V线圈, 1A触点",   "G6K-2F-Y(欧姆龙)"],
            ["高速缓冲U2/U3", "BUF634", "时序精确测量缓冲",
             "250mA, BW=180MHz",        "BUF634T(SMD版)"],
            ["匹配电阻R1-R4", "51Ω",   "输出阻抗匹配",
             "1%, 0402",                "49.9Ω/56Ω均可"],
            ["下拉电阻R5/R6", "1KΩ",   "输出终端下拉",
             "1%, 0402",                "820Ω~1.2KΩ均可"],
        ]
    elif "8281-4" in adapter_model:  # 模拟LDO适配器
        rows = [
            ["继电器K1-K3", "AGQ200A4H", "各路通断控制",
             "SPST, 5V线圈",             "G6K-2F-Y"],
            ["管脚驱动U2",  "EL7156",    "高性能驱动",
             "1.5A峰值, 30ns上升时间",   "EL7156C"],
            ["保护二极管U3","BAV99",      "ESD保护",
             "双肖特基, 70V, 200mA",     "BAV99LT1"],
            ["芯片插座U1",  "SOP8插座",  "承载DUT",
             "1.27mm间距, SOP8",         "IC-SOCKET-SOP8-5.4"],
        ]
    elif "8281-10" in adapter_model:  # 双工位适配器
        rows = [
            ["继电器K1/K2", "AGQ200A4H", "工位1/2通断控制",
             "SPST, 5V线圈",             "G6K-2F-Y"],
            ["工位插座U1/U2","ZIF14",    "双工位承载DUT",
             "14pin, 2.54mm间距",        "确认芯片封装"],
        ]

    return [header] + rows


def _apply_bom_styling(writer: pd.ExcelWriter) -> None:
    """BOM Excel样式"""
    sheet_colors = {
        "BOM清单":    "1A5276",
        "汇总信息":   "1E8449",
        "关键元件说明": "6E2FA0",
    }

    for sheet_name in writer.sheets:
        ws    = writer.sheets[sheet_name]
        color = sheet_colors.get(sheet_name, "1A5276")

        header_fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            )
            ws.column_dimensions[
                col[0].column_letter
            ].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"